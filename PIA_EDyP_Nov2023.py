import sqlite3
import sys
from sqlite3 import Error
import datetime
import csv
import openpyxl
import re
import pandas as pd

try:
    with sqlite3.connect('notas.db') as conn:
        mi_cursor = conn.cursor()
        
        mi_cursor.execute("CREATE TABLE IF NOT EXISTS clientes (id_cliente INTEGER PRIMARY KEY, \
        nombre_cliente TEXT NOT NULL, \
        RFC_cliente TEXT NOT NULL, \
        correo_cliente TEXT NOT NULL, \
        estado_cliente TEXT NOT NULL);")
        
        mi_cursor.execute("CREATE TABLE IF NOT EXISTS servicios (id_servicio INTEGER PRIMARY KEY, \
        nombre_servicio TEXT NOT NULL, \
        costo_servicio REAL NOT NULL, \
        estado_servicio TEXT NOT NULL);")
        
        mi_cursor.execute("CREATE TABLE IF NOT EXISTS notas (id_nota INTEGER PRIMARY KEY, \
        fecha_nota timestamp, \
        id_cliente INTEGER NOT NULL, \
        monto_a_pagar REAL NOT NULL, \
        estado_nota TEXT NOT NULL, \
        FOREIGN KEY(id_cliente) REFERENCES clientes(id_cliente));")
        
        mi_cursor.execute("CREATE TABLE IF NOT EXISTS detalles_notas (id_detalle INTEGER PRIMARY KEY, \
        id_nota INTEGER NOT NULL, \
        id_servicio INTEGER NOT NULL, \
        FOREIGN KEY(id_nota) REFERENCES notas(id_nota), \
        FOREIGN KEY(id_servicio) REFERENCES servicios(id_servicio));")  
except sqlite3.Error as e:
    print(e)
except:
    print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
finally:
    if (conn):
        conn.close()
        
def imprimir_nota(id_nota):
    with sqlite3.connect('notas.db', detect_types=sqlite3.PARSE_DECLTYPES | sqlite3.PARSE_COLNAMES) as conn:
        mi_cursor = conn.cursor()
        mi_cursor.execute("SELECT n.id_nota, n.fecha_nota, c.nombre_cliente, c.RFC_cliente, c.correo_cliente,n.monto_a_pagar, dn.id_servicio, s.nombre_servicio, s.costo_servicio\
            FROM notas n\
            INNER JOIN clientes c ON n.id_cliente = c.id_cliente\
            INNER JOIN detalles_notas dn ON n.id_nota = dn.id_nota\
            INNER JOIN servicios s ON dn.id_servicio = s.id_servicio\
            WHERE n.id_nota = ?", (id_nota,))

        resultados = mi_cursor.fetchall()

        id_nota, fecha_nota, nombre_cliente, RFC_cliente, correo_cliente, monto_a_pagar = resultados[0][:6]

        print(f"ID Nota: {id_nota}")
        print(f"Fecha de Nota: {fecha_nota.strftime('%d-%m-%Y')}")
        print(f"Cliente: {nombre_cliente}")
        print(f"RFC: {RFC_cliente}")
        print(f"Correo: {correo_cliente}")
        print("*" * 60)

        print(f"{'ID Servicio':<15} {'Nombre del Servicio':<25} {'Costo':<10}")
        print("-" * 50)

        for resultado in resultados:
            id_servicio, nombre_servicio, costo_servicio = resultado[6:]
            print(f"{id_servicio:<15} {nombre_servicio:<25} {costo_servicio:<10.2f}")

        print(f"Monto a Pagar: {monto_a_pagar:.2f}")
id_nota = 1
fecha_actual =datetime.datetime.now()
patron_fecha = r"^\d{2}-\d{2}-\d{4}$"

while True:
    print("""
    ╔══════════════════════╗
    ║   Menu Principal     ║
    ╠══════════════════════╣
    ║ 1. Notas             ║
    ║ 2. Clientes          ║
    ║ 3. Servicios         ║
    ║ 4. Estadísticas      ║ 
    ║ 5. Salir             ║ 
    ╚══════════════════════╝
    """)
    menu_principal = input("Ingrese una opción: ")
    if menu_principal == "1":
        while True:
            print("""
            ╔════════════════════════════╗
            ║           NOTAS            ║
            ╠════════════════════════════╢
            ║ 1. Registrar una nota      ║
            ║ 2. Cancelar una nota       ║
            ║ 3. Recuperar una nota      ║
            ║ 4. Consultas y reportes    ║
            ║ 5. Volver al menú principal║
            ╚════════════════════════════╝
            """)
            menu_notas = input("Ingrese una opción del menú de Notas: ")
            if menu_notas == "1":
                try:
                    with sqlite3.connect('notas.db') as conn:
                        mi_cursor = conn.cursor()
                        mi_cursor.execute("SELECT id_cliente, nombre_cliente FROM clientes WHERE estado_cliente = 'ACTIVO'")
                        clientes_registrados = mi_cursor.fetchall()
                        print("\nMenú de Clientes:")
                        if  clientes_registrados:
                            print("ID Cliente | Nombre Cliente")
                            print("-" * 26)  

                            for id_cliente, nombre_cliente in clientes_registrados:
                                id_cliente_str = str(id_cliente).rjust(10)
                                nombre_cliente = nombre_cliente.ljust(15)
                                print(f"{id_cliente_str} | {nombre_cliente}")
                            else:
                                print("*"*40)
                        else:
                            print("\nNO hay Clientes desponibles...\n")   
                        mi_cursor.execute("SELECT id_servicio, nombre_servicio, costo_servicio FROM servicios WHERE estado_servicio = 'ACTIVO'")
                        servicio_registrados = mi_cursor.fetchall()
                        print("\nMenú de Servicios:")
                        if servicio_registrados:
                            print(f"{'ID Servicio':<10} | {'Nombre de Servicio':<30} | {'Costo de Servicio':<15}")
                            print("-" * 70)

                            for id_servicio, nombre_servicio, costo_servicio in servicio_registrados:
                                id_servicio_str = str(id_servicio).rjust(11)
                                nombre_servicio = nombre_servicio.ljust(30)
                                costo_servicio = f'{costo_servicio:.2f}'.ljust(15)
                                print(f"{id_servicio_str} | {nombre_servicio} | {costo_servicio}")
                        else:
                            print("NO hay Servicios disponibles")
                        
                        patron_fecha = r"^\d{2}-\d{2}-\d{4}$"
                        
                        pregunta = input("\nDesea una nota? (ENTER PARA SALIR / [S] para seguir a delante)").strip().upper()
                        if pregunta =="":
                            break
                        
                        while True:
                            
                            fecha_ingresada_str = input("\nFecha de la nota (dd-mm-aaaa): ").strip()
                            
                            if not fecha_ingresada_str:
                                print("EL DATO NO PUEDE OMITIRSE. INTENTE DENUEVO.")
                            elif not re.match(patron_fecha, fecha_ingresada_str):
                                print("FORMATO DE FECHA INCORRECTO. DEBE SER DD-MM-AAAA")
                            else:
                                try:
                                    fecha_ingresada = datetime.datetime.strptime(fecha_ingresada_str, "%d-%m-%Y")
                                    if fecha_ingresada > fecha_actual:
                                        print("LA FECHA NO DEBE SER POSTERIOR A LA FECHA ACTUAL DEL SISTEMA")
                                    else:
                                        break
                                except ValueError:
                                    print("LA FECHA NO ES VÁLIDA/NO EXISTE. INTENTE DENUEVO.")
                        while True:
                            id_cliente = input("\nID del cliente: ").strip()
                            
                            if not id_cliente.isdigit():
                                print("El ID del cliente debe ser un número. Intente nuevamente.")
                            else:
                                mi_cursor.execute("SELECT nombre_cliente FROM clientes WHERE id_cliente = ? AND estado_cliente = 'ACTIVO'", (id_cliente,))
                                cliente_existente = mi_cursor.fetchone()
                                if cliente_existente:
                                    break
                                else:
                                    print("El cliente con ese ID no existe en la base de datos. Intente nuevamente.")
                        while True:
                            mi_cursor.execute("SELECT COUNT(*) FROM notas WHERE id_nota = ?", (id_nota,))
                            cuenta = mi_cursor.fetchone()[0]
                            if cuenta == 0:
                                break
                            else:
                                id_nota += 1
                                
                        estado = "ACTIVO"
                        monto_a_pagar = 0.0
                        cantidad_servicio = 0
                        while True:
                            id_servicio = input("\nIngrese el ID_servicio agregar / ( 0 para terminar la captura.): ").strip()
                            if not id_servicio:
                                print("EL DATO NO PUEDE OMITIRSE. INTENTE NUEVAMENTE.")
                                continue
                            elif id_servicio == "0":
                                if cantidad_servicio>0:
                                    break
                                else:
                                    print("La nota tiene que tener por lo menos un servicio. Intente nuevamente")
                            else:
                                mi_cursor.execute("SELECT nombre_servicio,costo_servicio FROM servicios WHERE id_servicio = ? AND estado_servicio = 'ACTIVO'", (id_servicio,))
                                servicio_encontrado = mi_cursor.fetchone()
                                if servicio_encontrado:
                                    cantidad_servicio+=1
                                    nombre_servicio, costo_servicio = servicio_encontrado
                                    monto_a_pagar += costo_servicio
                                    mi_cursor.execute("INSERT INTO detalles_notas (id_nota, id_servicio) VALUES (?, ?)", (id_nota, id_servicio))
                                    print("Servicio agregado con éxito")
                                    continue
                                else:
                                    print("El Servicio con ese ID no existe en la base de datos. Intente nuevamente.")
                        valores = (id_nota,fecha_ingresada,id_cliente,monto_a_pagar,estado)
                        mi_cursor.execute("INSERT INTO notas VALUES (?,?,?,?,?)",valores)
                        print("\nNota creada con éxito.")
                except sqlite3.Error as e:
                    print(e)
                except Exception as ex:
                    print(f"Se produjo el siguiente error: {ex}")
                imprimir_nota(id_nota)
            elif menu_notas == "2":
                while True:
                    folio_cancelar = input("\nIngrese el folio de la nota a cancelar/ 0 para ingresar al menú anterior: ").strip()
            
                    if folio_cancelar=="":
                        print("EL DATO NO PUEDE OMITIRSE. INTENTE DENUEVO.")
                        continue

                    if folio_cancelar=="0":
                        break
                    with sqlite3.connect('notas.db') as conn:
                        mi_cursor = conn.cursor()
                        mi_cursor.execute("SELECT estado_nota FROM notas WHERE id_nota = ? AND estado_nota = 'ACTIVO'", (folio_cancelar,))
                        estado_nota = mi_cursor.fetchone()
                        if estado_nota is not None:
                            estado_nota = estado_nota[0]
                        
                            if estado_nota == "ACTIVO":
                                imprimir_nota(folio_cancelar)
                                while True:
                                    respuesta = input("Seguro que quiere cancelar esta nota?: S/N").strip().upper()
                                    if respuesta == "S":
                                        mi_cursor.execute("UPDATE notas SET estado_nota = 'CANCELADO' WHERE id_nota  = ?", (folio_cancelar,))
                                        print("Nota cancelada con éxito.")
                                        break
                                    elif respuesta =="N":
                                        print(f"la nota: {folio_cancelar} no fue cancelada")
                                        break
                                    else:
                                        print("OPCIÓN NO VALIDA. INTENTE NUEVAMENTE.")
                                        continue
                            else:
                                print(f"No se encontró una nota con el folio {folio_cancelar}.")
                        else:
                            print(f"No se encontró una nota con el folio {folio_cancelar}.")
                    break 
            elif menu_notas == "3":
                try:
                    with sqlite3.connect('notas.db', detect_types=sqlite3.PARSE_DECLTYPES | sqlite3.PARSE_COLNAMES) as conn:
                        mi_cursor = conn.cursor()
  

                        mi_cursor.execute("SELECT n.id_nota, n.fecha_nota, c.nombre_cliente, c.RFC_cliente, c.correo_cliente, n.monto_a_pagar \
                                        FROM notas n \
                                        INNER JOIN clientes c ON n.id_cliente = c.id_cliente \
                                        WHERE n.estado_nota = 'CANCELADO'")

                        resultados = mi_cursor.fetchall()

                        if resultados:
                            print(f"{'ID Nota':<10}| {'Fecha_Nota':<12}| {'Cliente':<25}| {'RFC':<15}| {'Correo':<30}| {'Monto a Pagar':<15}")
                            print("*" * 110)

                            for id_nota, fecha_nota, nombre_cliente, RFC_cliente, correo_cliente, monto_a_pagar in resultados:
                                id_nota_str = str(id_nota).ljust(10)
                                fecha_nota_str = fecha_nota.strftime('%d-%m-%Y').ljust(12)
                                nombre_cliente_str = nombre_cliente.ljust(25)
                                RFC_cliente_str = RFC_cliente.ljust(15)
                                correo_cliente_str = correo_cliente.ljust(30)
                                monto_a_pagar_str = f"{monto_a_pagar:.2f}".ljust(15)
                                print(f"{id_nota_str}| {fecha_nota_str}| {nombre_cliente_str}| {RFC_cliente_str}| {correo_cliente_str}| {monto_a_pagar_str}")
                                print("-" * 110)
                        else:
                            print("No se encontraron notas canceladas.")
                            break
                except sqlite3.Error as e:
                    print(e)
                except Exception as ex:
                    print(f"Se produjo el siguiente error: {ex}")
                while True:
                    folio_recuperar = input("\nIngrese el folio de la nota a recuperar/ 0 para ingresar al menú principal: ").strip()
            
                    if folio_recuperar=="":
                        print("EL DATO NO PUEDE OMITIRSE. INTENTE DENUEVO.")
                        continue

                    if folio_recuperar=="0":
                        break
                    
                    with sqlite3.connect('notas.db') as conn:
                        mi_cursor = conn.cursor()
                        mi_cursor.execute("SELECT estado_nota FROM notas WHERE id_nota = ? AND estado_nota = 'CANCELADO' ", (folio_recuperar,))
                        estado_nota = mi_cursor.fetchone()
                        if estado_nota is not None:
                            estado_nota = estado_nota[0]
                        
                            if estado_nota == "CANCELADO":
                                imprimir_nota(folio_recuperar)
                                while True:
                                    respuesta = input("\nSeguro que quiere recuperar esta nota?: S/N").strip().upper()
                                    if respuesta == "S":
                                        mi_cursor.execute("UPDATE notas SET estado_nota = 'ACTIVO' WHERE id_nota = ?", (folio_recuperar,))
                                        print("Nota recuperada con éxito.")
                                        break
                                    elif respuesta =="N":
                                        print(f"la nota: {folio_recuperar} no fue recuperada")
                                        break
                                    else:
                                        print("OPCIÓN NO VALIDA. INTENTE NUEVAMENTE.")
                                        continue
                            else:
                                print(f"No se encontró una nota con el folio {folio_recuperar}.")
                                
                        else:
                            print(f"No se encontró una nota con el folio {folio_recuperar}.")
                            continue
            elif menu_notas == "4":
                while True:
                    print("""
                    ╔════════════════════════════╗
                    ║    CONSULTAS Y REPORTES    ║
                    ╟────────────────────────────╢
                    ║ 1. Consultar por período.  ║
                    ║ 2. Consultar por folio.    ║
                    ║ 3. Volver al menú de notas ║
                    ╚════════════════════════════╝
                    """)
                    menu_consulta_nota = input("Ingrese una opción del menú de consunsultas y reportes: ")
                    if menu_consulta_nota == "1":
                        try:
                            with sqlite3.connect('notas.db', detect_types=sqlite3.PARSE_DECLTYPES | sqlite3.PARSE_COLNAMES) as conn:
                                mi_cursor = conn.cursor()
                                
                                while True:
                                    print("\nDejar en blanco para usar 01-01-2000")
                                    fecha_inicial_str = input("Fecha inicial (dd-mm-aaaa):\n ").strip()
                                
                                    if fecha_inicial_str =="":
                                        print("Se utilizará la fecha por defecto: 01-01-2000.")
                                        fecha_inicial = datetime.datetime(2000, 1, 1)
                                        break
                                    elif not re.match(patron_fecha, fecha_inicial_str):
                                        print("FORMATO DE FECHA INCORRECTO. DEBE SER DD-MM-AAAA")
                                        continue
                                    try:
                                        fecha_inicial = datetime.datetime.strptime(fecha_inicial_str, "%d-%m-%Y")
                                    except Exception:
                                        print("LA FECHA NO EXISTE. INTENTE DENUEVO.")
                                        continue
                                    else:
                                        break
                                    
                                while True:
                                    print("\nDejar en blanco para usar la fecha actual del sistema.")       
                                    fecha_final_str = input("Fecha final (dd-mm-aaaa):\n ")
                                    
                                    if fecha_final_str =="":
                                        print("Se utilizará la fecha actual del sistema.")
                                        fecha_final=fecha_actual
                                        break
                                    elif not re.match(patron_fecha, fecha_final_str):
                                        print("FORMATO DE FECHA INCORRECTO. DEBE SER DD-MM-AAAA")
                                        continue
                                    
                                    try:
                                        fecha_final = datetime.datetime.strptime(fecha_final_str, "%d-%m-%Y")
                                        if not fecha_final>=fecha_inicial:
                                            print("LA FECHA FINAL DEBE SER IGUAL O POSTERIOR A LA FECHA INICIAL.INTENTE NUEVAMENTE.")
                                            continue
                                    except Exception:
                                        print("LA FECHA NO EXISTE. INTENTE DENUEVO.")
                                        continue
                                    else:
                                        break
                                mi_cursor.execute("SELECT n.id_nota, n.fecha_nota, c.nombre_cliente, c.RFC_cliente, c.correo_cliente, n.monto_a_pagar \
                                                FROM notas n \
                                                INNER JOIN clientes c ON n.id_cliente = c.id_cliente \
                                                WHERE n.estado_nota = 'ACTIVO' AND n.fecha_nota BETWEEN ? AND ?", (fecha_inicial, fecha_final))

                                resultados = mi_cursor.fetchall()

                                if resultados:
                                    print(f"{'ID Nota':<10}| {'Fecha_Nota':<12}| {'Cliente':<25}| {'RFC':<15}| {'Correo':<30}| {'Monto a Pagar':<15}")
                                    print("*" * 110)

                                    for id_nota, fecha_nota, nombre_cliente, RFC_cliente, correo_cliente, monto_a_pagar in resultados:
                                        id_nota_str = str(id_nota).ljust(10)
                                        fecha_nota_str = fecha_nota.strftime('%d-%m-%Y').ljust(12)
                                        nombre_cliente_str = nombre_cliente.ljust(25)
                                        RFC_cliente_str = RFC_cliente.ljust(15)
                                        correo_cliente_str = correo_cliente.ljust(30)
                                        monto_a_pagar_str = f"{monto_a_pagar:.2f}".ljust(15)
                                        print(f"{id_nota_str}| {fecha_nota_str}| {nombre_cliente_str}| {RFC_cliente_str}| {correo_cliente_str}| {monto_a_pagar_str}")
                                        print("-" * 110)
                                        
                                    while True:
                                        print("MENÚ\n[C]SV\n[E]xcel\n[R]egresar")
                                        opcion = input("¿Desea exportar el reporte? (CSV/Excel/Regresar): ").strip().lower()
                                        
                                        fecha_inicio_str = fecha_inicial.strftime("%d-%m-%Y")
                                        fecha_final_str = fecha_final.strftime("%d-%m-%Y")
        
                                        if opcion == "c":
                                            nombre_archivo = f"ReportePorPeriodo_{fecha_inicio_str}_{fecha_final_str}.csv"
                                            with open(nombre_archivo, 'w', newline='') as archivo_csv:
                                                escritor = csv.writer(archivo_csv)
                                                escritor.writerow(["ID Nota", "Fecha de Nota", "Cliente", "RFC", "Correo", "Monto a Pagar"])
                                                for resultado in resultados:
                                                    escritor.writerow(resultado)
                                                print(f'Se han guardado los datos en {nombre_archivo}')
                                                break
                                        elif opcion == "e":
                                            nombre_archivo = f"ReportePorPeriodo_{fecha_inicio_str}_{fecha_final_str}.xlsx"
                                            libro = openpyxl.Workbook()
                                            hoja = libro.active
                                            hoja.title = "Notas"
                                            hoja.append(["ID Nota", "Fecha Nota", "Nombre Cliente", "RFC Cliente", "Correo Cliente", "Monto a Pagar"])
                                            hoja.column_dimensions["A"].width = 10  
                                            hoja.column_dimensions["B"].width = 30  
                                            hoja.column_dimensions["C"].width = 30  
                                            hoja.column_dimensions["D"].width = 15
                                            hoja.column_dimensions["E"].width = 30
                                            hoja.column_dimensions["F"].width = 15
                                            for fila in resultados:
                                                hoja.append(fila)
                                            libro.save(nombre_archivo)
                                            print(f"Reporte exportado a {nombre_archivo}")
                                            break
                                        elif opcion == "r":
                                            break
                                        else:
                                            print("Opción no válida. Ingrese 'CSV', 'Excel' o 'Regresar'.")
                                else:
                                    print("No se encontraron notas.")
                                    break
                                break
                        except sqlite3.Error as e:
                            print(e)
                        except Exception as ex:
                            print(f"Se produjo el siguiente error: {ex}")
                    if menu_consulta_nota == "2":
                        try:
                            with sqlite3.connect('notas.db', detect_types=sqlite3.PARSE_DECLTYPES | sqlite3.PARSE_COLNAMES) as conn:
                                mi_cursor = conn.cursor()
        
                                mi_cursor.execute("SELECT n.id_nota, n.fecha_nota, c.nombre_cliente\
                                                    FROM notas n \
                                                    INNER JOIN clientes c ON n.id_cliente = c.id_cliente \
                                                    WHERE n.estado_nota = 'ACTIVO' \
                                                    ORDER BY n.id_nota ASC")
                                resultados = mi_cursor.fetchall()
                                if not resultados:
                                    print("No hay notas activas en el sistema.")
                                    break
                                else:
                                    df = pd.DataFrame(resultados, columns=["Clave", "Fecha", "Nombre"])
                                    df = df.set_index("Clave")
                                    print(df)
                                    
                                while True:   
                                    folio_consulta = input("\nIngrese el folio de la nota a consultar/(0 para salir): ").strip()
                                    
                                    if folio_consulta == "0":
                                        break
                                
                                    if folio_consulta == "":
                                        print("EL DATO NO PUEDE OMITIRSE. INTÉNTELO DE NUEVO.")
                                        continue

                                    try:
                                        folio_consulta = int(folio_consulta)
                                    except ValueError:
                                        print("CARÁCTER NO VÁLIDO. SOLO DÍGITOS NUMÉRICOS SON ACEPTADOS.")
                                        continue
                                    mi_cursor.execute("SELECT estado_nota FROM notas WHERE id_nota = ? AND estado_nota = 'ACTIVO'", (folio_consulta,))
                                    estado_nota = mi_cursor.fetchone()
                                    if estado_nota is not None:
                                        estado_nota = estado_nota[0]
                                    
                                        if estado_nota == "ACTIVO":
                                            imprimir_nota(folio_consulta)
                                            break
                                    else:
                                        print("Nota no se encuentra en el sistema")
                                        continue
                                break
                        except sqlite3.Error as e:
                            print(e)
                        except Exception as ex:
                            print(f"Se produjo el siguiente error: {ex}")
                    if menu_consulta_nota == "3":
                        print("Fuera del menú de consultas y reportes.")
                        break
                    else:
                        print("OPCIÓN NO VALIDA. INTENTE NUEVAMENTE.")
            elif menu_notas == "5":
                print("Fuera del menú de notas.")
                break
            else:
                print("OPCIÓN NO VALIDA. INTENTE NUEVAMENTE.")

    elif menu_principal == "2":
         while True:
            print("""
            ╔════════════════════════════╗
            ║         CLIENTES           ║
            ╠════════════════════════════╣
            ║ 1. Agregar un cliente.     ║
            ║ 2. Suspender a un cliente. ║
            ║ 3. Recuperar a un cliente. ║
            ║ 4. Consultas y reportes.   ║
            ║ 5. Volver al menú principal║
            ╚════════════════════════════╝
            """)
            menu_clientes = input("\nIngrese una opción del menú de clientes: ")

            if menu_clientes == "1":
                try:
                    with sqlite3.connect('notas.db') as conn:
                        mi_cursor = conn.cursor()
                        while True:
                            nombre_cliente = input("\nNombre del cliente([S] para salir): ").strip().upper()
    
                            if nombre_cliente == "":
                                print("EL DATO NO PUEDE OMITIRSE. INTENTE DENUEVO.")
                            elif any(char.isdigit() for char in nombre_cliente):
                                print("EL NOMBRE NO PUEDE CONTENER DÍGITOS. INTENTE NUEVAMENTE.")
                            elif nombre_cliente.lower() == "s":
                                break
                            else:
                                while True:
                                    RFC_cliente = input("\nIngrese un RFC (por ejemplo: Persona física: XEXT990101NI4 /Persona moral: EXT990101NI4 ): ").strip().upper()
                                    
                                    if not RFC_cliente:
                                        print("EL DATO NO PUEDE OMITIRSE. INTENTE DENUEVO.")
                                    elif not re.match(r'^[A-Z]{3,4}[0-9]{6}[A-Z0-9]{3}$', RFC_cliente):
                                        print("EL RFC INGRESADO NO TIENE EL FORMATO CORRECTO. INTENTE NUEVAMENTE.")
                                    else:
                                        try:
                                            if len(RFC_cliente) == 13:
                                                fecha_rfc = datetime.datetime.strptime(RFC_cliente[4:10], '%y%m%d')
                                            elif len(RFC_cliente) == 12:
                                                fecha_rfc = datetime.datetime.strptime(RFC_cliente[3:9], '%y%m%d')
                                        except ValueError:
                                            print("LA FECHA EN EL RFC NO ES VÁLIDA. INTENTE NUEVAMENTE.")
                                            continue
                                        break
                                while True:
                                    correo_cliente = input("\nIngrese su correo electrónico : ").strip()

                                    if not correo_cliente:
                                        print("EL DATO NO PUEDE OMITIRSE. INTENTE DENUEVO.")
                                    elif not re.match(r'^[a-zA-Z0-9._+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$', correo_cliente):
                                        print("EL CORREO ELECTRÓNICO TIENE UN FORMATO INCORRECTO/NO EXISTE. INTENTE NUEVAMENTE")  
                                    else:
                                        break
                            
                            estado_cliente = "ACTIVO"
                            valores = (nombre_cliente,RFC_cliente,correo_cliente,estado_cliente)
                            mi_cursor.execute("INSERT INTO clientes (nombre_cliente,RFC_cliente,correo_cliente,estado_cliente) VALUES(?,?,?,?)",valores)
                            print("Todo salio bien.")
                            break                
                except sqlite3.Error as e:
                    print(e)
                except:
                    print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                finally:
                    if (conn):
                        conn.close()

            elif menu_clientes == "2":
                        try:
                            with sqlite3.connect('notas.db') as conn:
                                mi_cursor = conn.cursor()
                                mi_cursor.execute("SELECT id_cliente, nombre_cliente FROM clientes WHERE estado_cliente = 'ACTIVO'")

                                todos_los_clientes = mi_cursor.fetchall()

                                if  todos_los_clientes:
                                    print("Todos los clientes registrados:")
                                    print(f"\n{'Clave':<10}| {'Nombre':<21}|")
                                    print("-"*35)
                                    for clave,nombre in todos_los_clientes:
                                        clave = str(clave)
                                        clave = clave.ljust(10)
                                        nombre =nombre.ljust(20)
                                        print(f"{clave:<10}| {nombre:<20}|")
                                        print("-"*35)
                                    while True: 
                                        id_cliente = input("Ingrese la clave del cliente que desea suspender (0 para cancelar): ").strip()

                                        if id_cliente == "0":
                                            print("\nNO SE SUSPENDIO UN CLIENTE\n")
                                            break
                                        elif id_cliente == "":
                                            print("NO SE PUEDE OMITIR EL DATO. INTENTE NUEVAMENTE")
                                            continue
                                        else:
                                            mi_cursor.execute("SELECT id_cliente, nombre_cliente,RFC_cliente,correo_cliente FROM clientes WHERE id_cliente = ? AND estado_cliente = 'ACTIVO'", (id_cliente,))
                                            cliente = mi_cursor.fetchone()

                                            if cliente:
                                                id_cliente, nombre_cliente,RFC_cliente,correo_cliente = cliente
                                                print(f"\nID del Cliente: {id_cliente}")
                                                print(f"Nombre del Cliente: {nombre_cliente}")
                                                print(f"RFC del Cliente: {RFC_cliente}")
                                                print(f"Correo del Cliente: {correo_cliente}\n")
                                                confirmacion = input("¿Desea suspender a este cliente? (S/N): ").strip().upper()
                                                if confirmacion == "S":
                                                    mi_cursor.execute("UPDATE clientes SET estado_cliente = 'SUSPENDIDO' WHERE id_cliente = ?", (id_cliente,))
                                                    print("El cliente ha sido suspendido.")
                                                    break
                                                elif confirmacion == "N":
                                                    print("\nNO SE SUSPENDIO EL CLIENTE.\n")
                                                    continue
                                                else:
                                                    print("EL DATO NO SE PUEDE OMITIR. INTENTE NUEVAMENTE.")
                                            else:
                                                print("CLIENTE NO ENCONTRADO EN EL SISTEMA. INTENTE NUEVAMENTE.")     
                                else:
                                    print("No hay clientes  para suspender.")
                                    continue
                                continue
                        except sqlite3.Error as e:
                            print(e)
                        except Exception as ex:
                            print(f"Se produjo el siguiente error: {ex}")            

    elif menu_principal == "3":
        pass
    elif menu_principal == "4":
        while True:
            print("""
            ╔════════════════════════════════════════╗
            ║           ESTADÍSTICAS                 ║
            ╟────────────────────────────────────────╢
            ║ 1. Servicios más prestados.            ║
            ║ 2. Clientes con más notas.             ║
            ║ 3. Promedio de los montos de las notas.║ 
            ║ 4. Volver al menú de servicios.        ║
            ╚════════════════════════════════════════╝
            """)
            menu_estadisticas = input("Ingrese una opción: ").strip()
            if menu_estadisticas == "1":
                while True:
                    cantidad = input("Ingrese la cantidad de servicios mas solicitados").strip()
                    try:
                        cantidad = int(cantidad)
                    except Exception:
                        print("EL DATO NO ES NUMÉRICO ENTERO.INTENTE NUEVAMENTE.")
                        continue
                    else:
                        if cantidad == 0:
                            print("LA CANTIDAD DE SERVIVIOS TIENE QUE SER MAYOR A 0. INTENTE NUEVAMENTE.")
                            continue
                        else:
                            break
                while True:
                    fecha_inicial_str = input("Fecha inicial (dd-mm-aaaa):\n ").strip()
                
                    if fecha_inicial_str =="":
                        print("NO SE PUEDE OMITIR EL DATO.")
                        continue
                    elif not re.match(patron_fecha, fecha_inicial_str):
                        print("FORMATO DE FECHA INCORRECTO. DEBE SER DD-MM-AAAA")
                        continue
                    try:
                        fecha_inicial = datetime.datetime.strptime(fecha_inicial_str, "%d-%m-%Y")
                    except Exception:
                        print("LA FECHA NO EXISTE. INTENTE DENUEVO.")
                        continue
                    else:
                        break
                while True:     
                    fecha_final_str = input("Fecha final (dd-mm-aaaa):\n ").strip()
                    
                    if fecha_final_str =="":
                        print("NO SE PUEDE OMITIR EL DATO.")
                    elif not re.match(patron_fecha, fecha_final_str):
                        print("FORMATO DE FECHA INCORRECTO. DEBE SER DD-MM-AAAA")
                        continue
                    try:
                        fecha_final = datetime.datetime.strptime(fecha_final_str, "%d-%m-%Y")
                        if not fecha_final>=fecha_inicial:
                            print("LA FECHA FINAL DEBE SER IGUAL O POSTERIOR A LA FECHA INICIAL.INTENTE NUEVAMENTE.")
                            continue
                    except Exception:
                        print("LA FECHA NO EXISTE. INTENTE DENUEVO.")
                        continue
                    else:
                        break
                try:
                    with sqlite3.connect('notas.db', detect_types=sqlite3.PARSE_DECLTYPES | sqlite3.PARSE_COLNAMES) as conn:
                        mi_cursor = conn.cursor()
                        mi_cursor.execute("SELECT s.nombre_servicio, COUNT(dn.id_detalle) AS servicios_prestados\
                        FROM servicios s\
                        INNER JOIN detalles_notas dn ON s.id_servicio = dn.id_servicio\
                        INNER JOIN notas n ON dn.id_nota = n.id_nota\
                        WHERE n.fecha_nota BETWEEN ? AND ?\
                        GROUP BY s.nombre_servicio\
                        ORDER BY servicios_prestados DESC\
                        LIMIT ?",(fecha_inicial,fecha_final,cantidad))
                        resultados = mi_cursor.fetchall()
                        if resultados:  
                            print(f"{'Nombre del Servicio':<30} {'Cantidad Prestada':<10}")
                            print("-" * 40)
                            for nombre, cantidad in resultados:
                                print(f"{nombre:<30} {cantidad:<10}")
                            while True:
                                formato= input("Ingrese el formato de exportación ([C]SV o [E]xcel /  (Enter para omitir)): ").upper()
                                if formato == 'C':
                                    nombre_archivo = f"ReporteServiciosMasPrestados_{fecha_inicial.strftime('%d-%m-%Y')}_{fecha_final.strftime('%d-%m-%Y')}.csv"
                                    try:
                                        with open(nombre_archivo, 'w', newline='') as archivo_csv:
                                            escritor = csv.writer(archivo_csv)
                                            escritor.writerow(["nombre_servicio", "Cantidad_Prestada"])
                                            for servicio, cantidad in resultados:
                                                escritor.writerow([servicio, cantidad])
                                        print(f'\nSe han guardado los datos en {nombre_archivo}\n')
                                        break
                                    except Exception as e:
                                        print(f'Error al guardar los datos en el archivo CSV: {e}')
                                elif formato == 'E':
                                    nombre_archivo = f"ReporteServiciosMasPrestados_{fecha_inicial.strftime('%d-%m-%Y')}_{fecha_final.strftime('%d-%m-%Y')}.xlsx"
                                    libro = openpyxl.Workbook()
                                    hoja = libro.active
                                    hoja.title = "Servicios"
                                    hoja.append(["Nombre del servicio", "Cantidad de servicio"])
                                    for nombre, cantidad in resultados:
                                        hoja.append([nombre,cantidad])
                                    hoja.column_dimensions["A"].width = 30
                                    hoja.column_dimensions["B"].width = 20
                                    libro.save(nombre_archivo)
                                    print(f"\nReporte exportado a {nombre_archivo}\n")
                                    break
                                elif formato == "":
                                    print("\nSe omitio la Exportación\n")
                                    break
                                else:
                                    print("Formato de exportación no válido. No se ha exportado el reporte.")
                                    continue
                        else:
                            print("NO HAY SERVICIOS REGISTRADOS")
                except sqlite3.Error as e:
                    print(e)
                except Exception as ex:
                    print(f"Se produjo el siguiente error: {ex}")
            elif menu_estadisticas == "2":
                while True:
                    cantidad = input("Ingrese cantidad de clientes con más notas a identificar").strip()
                    try:
                        cantidad = int(cantidad)
                    except Exception:
                        print("EL DATO NO ES NUMÉRICO ENTERO.INTENTE NUEVAMENTE.")
                        continue
                    else:
                        if cantidad == 0:
                            print("LA CANTIDAD DE CLIENTES TIENE QUE SER MAYOR A 0. INTENTE NUEVAMENTE.")
                            continue
                        else:
                            break
                while True:
                    fecha_inicial_str = input("Fecha inicial (dd-mm-aaaa):\n ").strip()
                
                    if fecha_inicial_str =="":
                        print("NO SE PUEDE OMITIR EL DATO.")
                        continue
                    elif not re.match(patron_fecha, fecha_inicial_str):
                        print("FORMATO DE FECHA INCORRECTO. DEBE SER DD-MM-AAAA")
                        continue
                    try:
                        fecha_inicial = datetime.datetime.strptime(fecha_inicial_str, "%d-%m-%Y")
                    except Exception:
                        print("LA FECHA NO EXISTE. INTENTE DENUEVO.")
                        continue
                    else:
                        break
                while True:     
                    fecha_final_str = input("Fecha final (dd-mm-aaaa):\n ").strip()
                    
                    if fecha_final_str =="":
                        print("NO SE PUEDE OMITIR EL DATO.")
                    elif not re.match(patron_fecha, fecha_final_str):
                        print("FORMATO DE FECHA INCORRECTO. DEBE SER DD-MM-AAAA")
                        continue
                    try:
                        fecha_final = datetime.datetime.strptime(fecha_final_str, "%d-%m-%Y")
                        if not fecha_final>=fecha_inicial:
                            print("LA FECHA FINAL DEBE SER IGUAL O POSTERIOR A LA FECHA INICIAL.INTENTE NUEVAMENTE.")
                            continue
                    except Exception:
                        print("LA FECHA NO EXISTE. INTENTE DENUEVO.")
                        continue
                    else:
                        break
                try:
                    with sqlite3.connect('notas.db', detect_types=sqlite3.PARSE_DECLTYPES | sqlite3.PARSE_COLNAMES) as conn:
                        mi_cursor = conn.cursor()
                        mi_cursor.execute("SELECT c.nombre_cliente, COUNT(n.id_nota) AS cantidad_notas\
                        FROM clientes c\
                        INNER JOIN notas n ON c.id_cliente = n.id_cliente\
                        WHERE n.fecha_nota BETWEEN ? AND ?\
                        GROUP BY c.nombre_cliente\
                        ORDER BY cantidad_notas DESC\
                        LIMIT ?",(fecha_inicial,fecha_final,cantidad))
                        resultados = mi_cursor.fetchall()
                        if resultados:  
                            print(f"{'Nombre del cliente':<30} {'Notas del cliente':<10}")
                            print("-" * 40)
                            for nombre, cantidad in resultados:
                                print(f"{nombre:<30} {cantidad:<10}")
                            while True:
                                formato= input("Ingrese el formato de exportación ([C]SV o [E]xcel /  (Enter para omitir)): ").upper()
                                if formato == 'C':
                                    nombre_archivo = f"ReporteClientesConMasNotas_{fecha_inicial.strftime('%d-%m-%Y')}_{fecha_final.strftime('%d-%m-%Y')}.csv"
                                    try:
                                        with open(nombre_archivo, 'w', newline='') as archivo_csv:
                                            escritor = csv.writer(archivo_csv)
                                            escritor.writerow(["Nombre_cliente", "Cantidad_nota"])
                                            for nombre, cantidad in resultados:
                                                escritor.writerow([nombre, cantidad])
                                        print(f'\nSe han guardado los datos en {nombre_archivo}\n')
                                        break
                                    except Exception as e:
                                        print(f'Error al guardar los datos en el archivo CSV: {e}')
                                elif formato == 'E':
                                    nombre_archivo = f"ReporteClientesConMasNotas_{fecha_inicial.strftime('%d-%m-%Y')}_{fecha_final.strftime('%d-%m-%Y')}.xlsx"
                                    libro = openpyxl.Workbook()
                                    hoja = libro.active
                                    hoja.title = "Clientes"
                                    hoja.append(["Nombre del cliente", "Cantidad de notas"])
                                    for nombre, cantidad in resultados:
                                        hoja.append([nombre,cantidad])
                                    hoja.column_dimensions["A"].width = 30
                                    hoja.column_dimensions["B"].width = 20
                                    libro.save(nombre_archivo)
                                    print(f"\nReporte exportado a {nombre_archivo}\n")
                                    break
                                elif formato == "":
                                    print("\nSe omitio la Exportación\n")
                                    break
                                else:
                                    print("Formato de exportación no válido. No se ha exportado el reporte.")
                                    continue
                        else:
                            print("NO HAY SERVICIOS REGISTRADOS")
                except sqlite3.Error as e:
                    print(e)
                except Exception as ex:
                    print(f"Se produjo el siguiente error: {ex}")
            elif menu_estadisticas == "3":
                while True:
                    fecha_inicial_str = input("Fecha inicial (dd-mm-aaaa):\n ").strip()
                
                    if fecha_inicial_str =="":
                        print("NO SE PUEDE OMITIR EL DATO.")
                        continue
                    elif not re.match(patron_fecha, fecha_inicial_str):
                        print("FORMATO DE FECHA INCORRECTO. DEBE SER DD-MM-AAAA")
                        continue
                    try:
                        fecha_inicial = datetime.datetime.strptime(fecha_inicial_str, "%d-%m-%Y")
                    except Exception:
                        print("LA FECHA NO EXISTE. INTENTE DENUEVO.")
                        continue
                    else:
                        break
                while True:     
                    fecha_final_str = input("Fecha final (dd-mm-aaaa):\n ").strip()
                    
                    if fecha_final_str =="":
                        print("NO SE PUEDE OMITIR EL DATO.")
                    elif not re.match(patron_fecha, fecha_final_str):
                        print("FORMATO DE FECHA INCORRECTO. DEBE SER DD-MM-AAAA")
                        continue
                    try:
                        fecha_final = datetime.datetime.strptime(fecha_final_str, "%d-%m-%Y")
                        if not fecha_final>=fecha_inicial:
                            print("LA FECHA FINAL DEBE SER IGUAL O POSTERIOR A LA FECHA INICIAL.INTENTE NUEVAMENTE.")
                            continue
                    except Exception:
                        print("LA FECHA NO EXISTE. INTENTE DENUEVO.")
                        continue
                    else:
                        break
                try:
                    with sqlite3.connect('notas.db', detect_types=sqlite3.PARSE_DECLTYPES | sqlite3.PARSE_COLNAMES) as conn:
                        mi_cursor = conn.cursor()
                        mi_cursor.execute("SELECT AVG(monto_a_pagar) FROM notas \
                        WHERE fecha_nota BETWEEN ? AND ?",(fecha_inicial,fecha_final))
                        resultado = mi_cursor.fetchone()
                        if resultado is not None and resultado[0] is not None:
                            promedio = resultado[0]
                            print("-" * 35)
                            print(f"Promedio de las notas\nPeríodo: {fecha_inicial.strftime('%d-%m-%Y')}----{fecha_final.strftime('%d-%m-%Y')}:")
                            print(f"{promedio:>15.2f}\nMoneda: Pesos")
                            print("-" * 35)
                        else:
                            print(f"No se encontraron montos de notas con los cuales sacar un promedio en el período:\n{fecha_inicial.strftime('%d-%m-%Y')}----{fecha_final.strftime('%d-%m-%Y')}")
                except sqlite3.Error as e:
                    print(e)
                except Exception as ex:
                    print(f"Se produjo el siguiente error: {ex}")
            elif menu_estadisticas == "4":
                print("\nFuera del menu de estadísticas...\n")
                break
    elif menu_principal == "5":
        pass
    else:
        print("OPCIÓN NO VALIDA. INTENTE NUEVAMENTE.")